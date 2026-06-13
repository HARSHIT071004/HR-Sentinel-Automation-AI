import random
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def generate_excel():
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"

    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    late_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    present_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # Headers
    headers = ["Employee ID", "Employee Name", "Department", "Date", "Time", "Status"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    # 60 employees across 5 departments
    departments = ["Engineering", "Product", "Design", "Marketing", "HR", "Finance"]
    employees = [
        ("EMP001", "Aarav Sharma", "Engineering"), ("EMP002", "Ishita Kapoor", "Engineering"),
        ("EMP003", "Rohan Verma", "Engineering"), ("EMP004", "Sanya Gupta", "Engineering"),
        ("EMP005", "Vikram Singh", "Engineering"), ("EMP006", "Priya Patel", "Product"),
        ("EMP007", "Arjun Mehta", "Product"), ("EMP008", "Ananya Reddy", "Product"),
        ("EMP009", "Karan Malhotra", "Product"), ("EMP010", "Neha Joshi", "Product"),
        ("EMP011", "Rahul Bose", "Design"), ("EMP012", "Deepika Nair", "Design"),
        ("EMP013", "Aditya Kulkarni", "Design"), ("EMP014", "Sneha Iyer", "Design"),
        ("EMP015", "Varun Saxena", "Design"), ("EMP016", "Pooja Sharma", "Marketing"),
        ("EMP017", "Nikhil Agarwal", "Marketing"), ("EMP018", "Megha Rao", "Marketing"),
        ("EMP019", "Siddharth Jain", "Marketing"), ("EMP020", "Ritu Chandra", "Marketing"),
        ("EMP021", "Amit Tiwari", "HR"), ("EMP022", "Shruti Menon", "HR"),
        ("EMP023", "Gaurav Pandey", "HR"), ("EMP024", "Tanvi Desai", "Finance"),
        ("EMP025", "Manish Kumar", "Finance"), ("EMP026", "Kavita Sinha", "Finance"),
        ("EMP027", "Rajesh Gupta", "Engineering"), ("EMP028", "Swati Mishra", "Engineering"),
        ("EMP029", "Ajay Bhatia", "Product"), ("EMP030", "Deepa Pillai", "Product"),
        ("EMP031", "Sanjay Kulkarni", "Design"), ("EMP032", "Manju Sharma", "Marketing"),
        ("EMP033", "Vishal Rao", "Finance"), ("EMP034", "Divya Kapoor", "Engineering"),
        ("EMP035", "Ashish Verma", "Engineering"), ("EMP036", "Pallavi Joshi", "Product"),
        ("EMP037", "Suresh Nair", "Design"), ("EMP038", "Geeta Reddy", "Marketing"),
        ("EMP039", "Ramesh Tiwari", "Finance"), ("EMP040", "Usha Bhat", "HR"),
        ("EMP041", "Prakash Singh", "Engineering"), ("EMP042", "Sunita Devi", "Engineering"),
        ("EMP043", "Mohan Das", "Product"), ("EMP044", "Sarita Gupta", "Design"),
        ("EMP045", "Rakesh Patel", "Marketing"), ("EMP046", "Asha Rani", "Finance"),
        ("EMP047", "Vinod Sharma", "Engineering"), ("EMP048", "Kamla Devi", "HR"),
        ("EMP049", "Brijesh Kumar", "Finance"), ("EMP050", "Indu Bala", "Product"),
        ("EMP051", "Deepak Joshi", "Engineering"), ("EMP052", "Meena Kumari", "Product"),
        ("EMP053", "Suresh Reddy", "Design"), ("EMP054", "Lata Sharma", "Marketing"),
        ("EMP055", "Ashok Kumar", "Finance"), ("EMP056", "Sudha Patel", "HR"),
        ("EMP057", "Ravi Shankar", "Engineering"), ("EMP058", "Urmila Devi", "Product"),
        ("EMP059", "Gopal Krishna", "Design"), ("EMP060", "Savitri Bai", "Finance"),
    ]

    # Employees who are consistently late (more violations)
    late_employees = {"EMP002", "EMP004", "EMP015", "EMP022", "EMP035", "EMP042", "EMP051", "EMP055"}

    # Employees with occasional issues
    occasional_late = {"EMP009", "EMP017", "EMP029", "EMP038", "EMP047"}

    row = 2
    today = datetime.now()

    for emp_id, name, dept in employees:
        for day_offset in range(14):  # 2 weeks of data
            date = today - timedelta(days=day_offset)
            if date.weekday() >= 5:  # Skip weekends
                continue

            # Determine attendance pattern
            if emp_id in late_employees:
                # Consistently late
                checkin_hour = random.choice([11, 12, 13])
                checkin_min = random.randint(0, 59)
                # Sometimes missing checkout
                has_checkout = random.random() > 0.15
            elif emp_id in occasional_late:
                # Occasionally late
                is_late_today = random.random() < 0.3
                if is_late_today:
                    checkin_hour = random.choice([11, 12])
                    checkin_min = random.randint(0, 59)
                else:
                    checkin_hour = random.randint(8, 10)
                    checkin_min = random.randint(0, 59)
                has_checkout = random.random() > 0.05
            else:
                # Usually on time
                checkin_hour = random.randint(8, 10)
                checkin_min = random.randint(0, 59)
                has_checkout = random.random() > 0.02

            checkin_time = f"{checkin_hour:02d}:{checkin_min:02d}:00"

            # IN punch
            ws.cell(row=row, column=1, value=emp_id).border = thin_border
            ws.cell(row=row, column=2, value=name).border = thin_border
            ws.cell(row=row, column=3, value=dept).border = thin_border
            ws.cell(row=row, column=4, value=date.strftime("%Y-%m-%d")).border = thin_border
            ws.cell(row=row, column=5, value=checkin_time).border = thin_border
            status_cell = ws.cell(row=row, column=6, value="IN")
            status_cell.border = thin_border
            if checkin_hour >= 11:
                status_cell.fill = late_fill
            else:
                status_cell.fill = present_fill
            row += 1

            # OUT punch (if not missing)
            if has_checkout:
                checkout_hour = random.randint(17, 19)
                checkout_min = random.randint(0, 59)
                checkout_time = f"{checkout_hour:02d}:{checkout_min:02d}:00"

                ws.cell(row=row, column=1, value=emp_id).border = thin_border
                ws.cell(row=row, column=2, value=name).border = thin_border
                ws.cell(row=row, column=3, value=dept).border = thin_border
                ws.cell(row=row, column=4, value=date.strftime("%Y-%m-%d")).border = thin_border
                ws.cell(row=row, column=5, value=checkout_time).border = thin_border
                ws.cell(row=row, column=6, value="OUT").border = thin_border
                row += 1

    # Auto-fit columns
    for col in ws.columns:
        max_length = 0
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col[0].column_letter].width = max_length + 2

    # Add summary sheet
    ws2 = wb.create_sheet("Employee List")
    summary_headers = ["Employee ID", "Name", "Department", "Expected Late Count"]
    for col, header in enumerate(summary_headers, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    for i, (emp_id, name, dept) in enumerate(employees, 2):
        late_count = 8 if emp_id in late_employees else (3 if emp_id in occasional_late else 0)
        ws2.cell(row=i, column=1, value=emp_id).border = thin_border
        ws2.cell(row=i, column=2, value=name).border = thin_border
        ws2.cell(row=i, column=3, value=dept).border = thin_border
        ws2.cell(row=i, column=4, value=late_count).border = thin_border

    for col in ws2.columns:
        max_length = 0
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws2.column_dimensions[col[0].column_letter].width = max_length + 2

    filename = "attendance_60_employees.xlsx"
    wb.save(filename)
    print(f"Created: {filename}")
    print(f"Total rows: {row - 2}")
    print(f"Employees: {len(employees)}")
    print(f"Late employees (consistent): {len(late_employees)}")
    print(f"Late employees (occasional): {len(occasional_late)}")
    print(f"Date range: {(today - timedelta(days=13)).strftime('%Y-%m-%d')} to {today.strftime('%Y-%m-%d')}")


if __name__ == "__main__":
    generate_excel()
