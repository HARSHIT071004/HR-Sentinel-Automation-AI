import sys
sys.path.insert(0, '.')

import random
from datetime import datetime, timedelta
from openpyxl import Workbook


def generate_excel():
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"

    # Headers
    headers = ["employee_id", "name", "date", "time", "status"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

    # 50 employees
    employees = [
        ("EMP001", "Aarav Sharma"), ("EMP002", "Ishita Kapoor"), ("EMP003", "Rohan Verma"),
        ("EMP004", "Sanya Gupta"), ("EMP005", "Vikram Singh"), ("EMP006", "Priya Patel"),
        ("EMP007", "Arjun Mehta"), ("EMP008", "Ananya Reddy"), ("EMP009", "Karan Malhotra"),
        ("EMP010", "Neha Joshi"), ("EMP011", "Rahul Bose"), ("EMP012", "Deepika Nair"),
        ("EMP013", "Aditya Kulkarni"), ("EMP014", "Sneha Iyer"), ("EMP015", "Varun Saxena"),
        ("EMP016", "Pooja Sharma"), ("EMP017", "Nikhil Agarwal"), ("EMP018", "Megha Rao"),
        ("EMP019", "Siddharth Jain"), ("EMP020", "Ritu Chandra"), ("EMP021", "Amit Tiwari"),
        ("EMP022", "Shruti Menon"), ("EMP023", "Gaurav Pandey"), ("EMP024", "Tanvi Desai"),
        ("EMP025", "Manish Kumar"), ("EMP026", "Kavita Sinha"), ("EMP027", "Rajesh Gupta"),
        ("EMP028", "Swati Mishra"), ("EMP029", "Ajay Bhatia"), ("EMP030", "Deepa Pillai"),
        ("EMP031", "Sanjay Kulkarni"), ("EMP032", "Manju Sharma"), ("EMP033", "Vishal Rao"),
        ("EMP034", "Divya Kapoor"), ("EMP035", "Ashish Verma"), ("EMP036", "Pallavi Joshi"),
        ("EMP037", "Suresh Nair"), ("EMP038", "Geeta Reddy"), ("EMP039", "Ramesh Tiwari"),
        ("EMP040", "Usha Bhat"), ("EMP041", "Prakash Singh"), ("EMP042", "Sunita Devi"),
        ("EMP043", "Mohan Das"), ("EMP044", "Sarita Gupta"), ("EMP045", "Rakesh Patel"),
        ("EMP046", "Asha Rani"), ("EMP047", "Vinod Sharma"), ("EMP048", "Kamla Devi"),
        ("EMP049", "Brijesh Kumar"), ("EMP050", "Indu Bala"),
    ]

    row = 2
    today = datetime.now()

    for emp_id, name in employees:
        # Generate 5 days of attendance
        for day_offset in range(5):
            date = today - timedelta(days=day_offset)
            if date.weekday() >= 5:  # Skip weekends
                continue

            # Determine if late (some employees are consistently late)
            is_late = emp_id in ["EMP002", "EMP004", "EMP015", "EMP022", "EMP035", "EMP042"]

            if is_late:
                checkin_hour = random.choice([11, 12])
                checkin_min = random.randint(0, 59)
            else:
                checkin_hour = random.randint(8, 10)
                checkin_min = random.randint(0, 59)

            checkin_time = f"{checkin_hour:02d}:{checkin_min:02d}:00"
            checkout_time = f"{random.randint(17, 19):02d}:{random.randint(0, 59):02d}:00"

            # IN punch
            ws.cell(row=row, column=1, value=emp_id)
            ws.cell(row=row, column=2, value=name)
            ws.cell(row=row, column=3, value=date.strftime("%Y-%m-%d"))
            ws.cell(row=row, column=4, value=checkin_time)
            ws.cell(row=row, column=5, value="IN")
            row += 1

            # OUT punch
            ws.cell(row=row, column=1, value=emp_id)
            ws.cell(row=row, column=2, value=name)
            ws.cell(row=row, column=3, value=date.strftime("%Y-%m-%d"))
            ws.cell(row=row, column=4, value=checkout_time)
            ws.cell(row=row, column=5, value="OUT")
            row += 1

    # Auto-fit columns
    for col in ws.columns:
        max_length = 0
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col[0].column_letter].width = max_length + 2

    filename = "attendance_50_employees.xlsx"
    wb.save(filename)
    print(f"Created {filename} with {row - 2} rows for {len(employees)} employees")
    print(f"Late employees: EMP002, EMP004, EMP015, EMP022, EMP035, EMP042")


if __name__ == "__main__":
    generate_excel()
